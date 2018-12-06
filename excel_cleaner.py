import datetime as dt
import openpyxl as xl
import pandas as pd
from os import remove


def clean_sheet(wb: xl.Workbook, sheet_name: str)->(dt.datetime, str):
    """
    :param wb: OpenPyXL workbook
    :param sheet_name: String representation of the sheet
    :return: the name of the sheet, and the datetime object
    Takes an OpenPyXL workbook and sheet representation and cleans the data for processing
    Returns the datetime object and sheet name to be later processed in Pandas
    """

    date_cell = 'C5'  # cell that contains the date of the report
    worksheet = wb[sheet_name]  # worksheet to be cleaned

    date = worksheet[date_cell].value  # date of the report - should be a datetime.datetime object

    worksheet.unmerge_cells('C2:F2')  # un-merge the title cells
    worksheet.unmerge_cells('C3:F4')  # un-merge the title cells

    worksheet.delete_rows(15, 16)  # delete the in-between rows
    worksheet.delete_rows(1, 13)  # delete the title rows

    worksheet.delete_cols(1)  # delete the bumper column

    return date, sheet_name


def temp_file_to_dataframe(file_name: str, date_sheet_pairs: list)->pd.DataFrame:

    df_list = []

    for _date, sheet_name in date_sheet_pairs:
        temp_df = pd.read_excel(file_name, sheet_name=sheet_name)
        temp_df['Period'] = _date
        df_list.append(temp_df)

    df_ = pd.concat(df_list)
    df_.dropna(subset=['Provider Name'], inplace=True)
    df_.iloc[:, 3:12] = df_.iloc[:, 3:12].apply(pd.to_numeric, errors='coerce')
    return df_


def map_stp(dataframe: pd.DataFrame, stp_map: str)->pd.DataFrame:

    map_df = pd.read_excel(stp_map)
    temp_df = pd.merge(dataframe, map_df, left_on='Org Code', right_on='NHS ID code', how='inner')
    return temp_df


if __name__ == '__main__':

    # Load the DID spreadsheet from NHS Digital

    file = xl.load_workbook('data\\Tables-1a-1l-2017-18-Modality-Provider-Counts-XLSX-222KB (1).xlsx')

    # Initialise a list to contain the datetime/sheet_name tuples

    ds_pairs = []

    # loop through the sheets (ignoring the title sheet) and clean the sheet using clean_sheet()

    for sheet in file.sheetnames[1:]:
        ds_pairs.append(clean_sheet(file, sheet))

    # save the file as a temp file

    file.save('data\\temp_file.xlsx')

    # Close file

    file.close()

    df = temp_file_to_dataframe('data\\temp_file.xlsx', ds_pairs)
    df = map_stp(df, 'data\\STP Map.xlsx')

    df.to_excel(f'data\\df {dt.date.today()}.xlsx')

    # delete the temporary excel file

    remove('data\\temp_file.xlsx')
